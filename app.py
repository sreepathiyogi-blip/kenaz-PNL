import streamlit as st
import pandas as pd
import numpy as np
import gspread
from google.oauth2.service_account import Credentials
from pyxlsb import open_workbook
from datetime import date, timedelta
import plotly.graph_objects as go
import plotly.express as px
import tempfile, os, io
import warnings
warnings.filterwarnings("ignore")

st.set_page_config(
    page_title="Kenaz P&L Dashboard", page_icon="🕌", layout="wide",
    initial_sidebar_state="expanded",
    menu_items={"Get Help": None, "Report a bug": None, "About": "One Guardian Brands — Kenaz P&L Dashboard"},
)

GOLD  = "#C9A84C"
CREAM = "#FFFFF8"
DARK  = "#1A1A1A"
CARD  = "#2B2B2B"

st.markdown(f"""
<style>
  [data-testid="stAppViewContainer"] {{ background-color:{DARK}; }}
  [data-testid="stSidebar"] {{ background-color:#222; }}
  .block-container {{ padding-top:1.5rem; }}
  .kpi-card {{ background:{CARD};border:1px solid {GOLD}33;border-radius:10px;padding:16px 20px;margin-bottom:8px; }}
  .kpi-label {{ color:#aaa;font-size:11px;letter-spacing:1px;text-transform:uppercase; }}
  .kpi-value {{ color:{GOLD};font-size:26px;font-weight:700;margin:4px 0; }}
  .kpi-sub   {{ color:#777;font-size:12px; }}
  .pnl-table {{ width:100%;border-collapse:collapse;font-size:12.5px; }}
  .pnl-table th {{ background:{GOLD}22;color:{GOLD};padding:8px 12px;text-align:right;
                   border-bottom:2px solid {GOLD}55;white-space:nowrap; }}
  .pnl-table th:first-child {{ text-align:left;min-width:220px; }}
  .pnl-table td {{ padding:6px 12px;text-align:right;border-bottom:1px solid #2a2a2a;color:#ddd;white-space:nowrap; }}
  .pnl-table td:first-child {{ text-align:left;color:#bbb; }}
  .pnl-table tr.total-row td {{ background:{GOLD}18;font-weight:700;color:{CREAM};border-top:2px solid {GOLD}55;border-bottom:2px solid {GOLD}33; }}
  .pnl-table tr.pct-row td {{ color:#888;font-size:11.5px;border-bottom:1px solid #222; }}
  .pnl-table tr.pct-row td:first-child {{ color:#666; }}
  .pnl-table tr.section-gap td {{ height:6px;background:{DARK};border:none; }}
  .pnl-table tr:not(.pct-row):not(.total-row):not(.section-gap):hover td {{ background:#252525; }}
  .positive {{ color:#4caf50!important; }}
  .negative {{ color:#e57373!important; }}
  .header-bar {{ background:{GOLD};padding:12px 20px;border-radius:8px;color:#1A1A1A;
                 font-weight:800;font-size:20px;letter-spacing:1.5px;margin-bottom:20px; }}
  h1,h2,h3 {{ color:{GOLD}!important; }}
  label {{ color:#aaa!important; }}
  .stRadio label {{ color:#ccc!important; }}
</style>
""", unsafe_allow_html=True)

SCOPES    = ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive"]
SHEET_KEY = "10qFitbppdVbNK0w67q1HFK-l7N1uAzHJ0mkyB2XImJQ"
CHANNELS  = ["Website","Amazon","Meesho","Flipkart","Myntra PPMP"]
EAN_MAP   = {
    8906188065836:"Triumph",      8906188065799:"Gentleman",
    8906188065928:"Oud Ameer",    8906188065775:"Untamed",
    8906188065867:"Fortuna",      8906188065850:"La Beaute",
    8906188065881:"Twilight",     8906188065980:"Bahiyaa Bayda",
    8904512100307:"Female Gift Set", 8904512100291:"Male Gift Set",
}
SKU_NUM_COLS = ["Revenue Without Tax","Qty","COGS","TOTAL MRP","Inward","Wages",
                "commission","Payment Gateway","Shipping","Bulk Logistic Cost",
                "Packaging Cost","Warehousing Charges","Rebate","others","Total Spend"]
PNL_COLS  = ["Month_serial","Month_name","Channel","MRP Sales","Quantity","Net Sales",
             "COGS","Freight Inward","Wages","Commission","Payment Gateway","Shipping",
             "Others","Ad Spend","Bulk Logistic","Packaging","Warehousing","Rebate"]

def xlsb_to_date(n):
    try: return date(1899,12,30)+timedelta(days=int(float(n)))
    except: return None

def L(v):
    try: f=float(v)
    except: return "-"
    if pd.isna(f) or f==0: return "-"
    return f"{f:,.2f}"

def Lbold(v):
    try: f=float(v)
    except: return "-"
    if pd.isna(f): return "-"
    s=L(abs(f))
    if s=="-": return "-"
    return f"({s})" if f<0 else s

def P(v):
    try: f=float(v)
    except: return "-"
    if pd.isna(f): return "-"
    r=round(f,1)
    return f"{r:.1f}%"

def INR(v):
    try: f=float(v)
    except: return "-"
    if pd.isna(f) or f==0: return "-"
    return f"&#8377;{f:,.0f}"

def color_val(v,inverse=False):
    try: f=float(v)
    except: return "#aaa"
    if pd.isna(f) or f==0: return "#aaa"
    pos=f>0
    if inverse: pos=not pos
    return "#4caf50" if pos else "#e57373"

def safe_div(num,den):
    try:
        n,d=float(num),float(den)
        if pd.isna(n) or pd.isna(d) or d==0: return np.nan
        return n/d
    except: return np.nan

# ─── Google Sheets ─────────────────────────────────────────────────────────────
@st.cache_resource
def get_gsheet_client():
    creds=Credentials.from_service_account_info(st.secrets["gcp_service_account"],scopes=SCOPES)
    return gspread.authorize(creds)

def get_sheet(client): return client.open_by_key(SHEET_KEY)

@st.cache_data(ttl=300)
def load_from_gsheet(_dummy="") -> pd.DataFrame:
    try:
        client=get_gsheet_client()
        ws=get_sheet(client).sheet1
        data=ws.get_all_records()
        if not data: return pd.DataFrame(columns=PNL_COLS)
        df=pd.DataFrame(data)
        for col in ["MRP Sales","Quantity","Net Sales","COGS","Freight Inward","Wages",
                    "Commission","Payment Gateway","Shipping","Others","Ad Spend",
                    "Bulk Logistic","Packaging","Warehousing","Rebate"]:
            if col in df.columns:
                df[col]=pd.to_numeric(df[col],errors="coerce").fillna(0)
        return df
    except Exception as e:
        st.error(f"Sheet load error: {e}")
        return pd.DataFrame(columns=PNL_COLS)

@st.cache_data(ttl=300)
def load_expenses_from_gsheet(_dummy="") -> pd.DataFrame:
    try:
        client=get_gsheet_client()
        sh=get_sheet(client)
        try: ws=sh.worksheet("Expenses")
        except gspread.exceptions.WorksheetNotFound: return pd.DataFrame()
        data=ws.get_all_records()
        if not data: return pd.DataFrame()
        df=pd.DataFrame(data)
        df["Amount"]=pd.to_numeric(df["Amount"],errors="coerce").fillna(0)
        return df
    except Exception as e:
        st.error(f"Expense sheet load error: {e}")
        return pd.DataFrame()

def save_to_gsheet(client, new_df: pd.DataFrame):
    sh=get_sheet(client)
    ws=sh.sheet1
    existing=ws.get_all_records()
    new_df=new_df.copy()
    new_df["Month_serial"]=new_df["Month_serial"].astype(str)
    NUM_COLS=["MRP Sales","Quantity","Net Sales","COGS","Freight Inward","Wages",
              "Commission","Payment Gateway","Shipping","Others","Ad Spend",
              "Bulk Logistic","Packaging","Warehousing","Rebate"]
    def df_to_rows(df):
        out=df.copy()
        for c in NUM_COLS:
            if c in out.columns:
                out[c]=pd.to_numeric(out[c],errors="coerce").fillna(0).round(4)
        return [out.columns.tolist()]+out.astype(str).values.tolist()
    if not existing:
        ws.update(df_to_rows(new_df))
        return len(new_df),0
    ex=pd.DataFrame(existing)
    ex["Month_serial"]=ex["Month_serial"].astype(str)
    ex_keys=ex["Month_serial"]+"_"+ex["Channel"].astype(str)
    new_keys=new_df["Month_serial"]+"_"+new_df["Channel"].astype(str)
    truly_new=new_df[~new_keys.isin(ex_keys)]
    if len(truly_new)==0: return 0,len(new_df)
    all_cols=list(dict.fromkeys(ex.columns.tolist()+truly_new.columns.tolist()))
    combined=pd.concat([ex.reindex(columns=all_cols,fill_value=""),
                        truly_new.reindex(columns=all_cols,fill_value="")],ignore_index=True)
    combined=combined.sort_values(["Month_serial","Channel"])
    ws.clear()
    ws.update(df_to_rows(combined))
    return len(truly_new),len(new_df)-len(truly_new)

def write_pnl_sheet(client, df_all, exp_df=None):
    """Write a beautifully formatted P&L summary sheet to Google Sheets tab 'P&L Summary'."""
    sh = get_sheet(client)

    # get or create the P&L Summary worksheet
    try:
        ws = sh.worksheet("P&L Summary")
        ws.clear()
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title="P&L Summary", rows=200, cols=30)

    METRICS = ["MRP Sales","Quantity","Net Sales","COGS","Freight Inward","Wages",
               "Commission","Payment Gateway","Shipping","Bulk Logistic","Packaging",
               "Warehousing","Rebate","Others","Ad Spend"]

    grp = df_all.groupby(["Month_name","Month_sort"])[METRICS].sum().reset_index()
    grp = grp.sort_values("Month_sort")
    months = grp["Month_name"].tolist()

    lookup = {row["Month_name"]:{m:float(row[m]) for m in METRICS} for _,row in grp.iterrows()}

    def v(metric,m): return lookup.get(m,{}).get(metric,0)
    def nsv(m): return v("Net Sales",m) or np.nan
    def mat(m): return v("Net Sales",m)-v("COGS",m)
    def fw(m):  return v("Freight Inward",m)+v("Wages",m)
    def gm(m):  return mat(m)-fw(m)
    def cnl(m): return sum(v(c,m) for c in ["Commission","Payment Gateway","Shipping",
                                             "Bulk Logistic","Packaging","Warehousing","Rebate","Others"])
    def cm1(m): return gm(m)-cnl(m)
    def cm2(m): return cm1(m)-v("Ad Spend",m)

    tot_ns = sum(v("Net Sales",m) for m in months) or np.nan
    tot_qty = sum(v("Quantity",m) for m in months)

    def brand_mkt(m):
        if exp_df is None or exp_df.empty: return 0
        return float(exp_df[exp_df["Month"]==m]["Amount"].sum())
    def cm3(m): return cm2(m)-brand_mkt(m)

    def lacs(x):
        try:
            f=float(x)
            if pd.isna(f) or f==0: return "-"
            return f"{f/1e5:,.2f}"
        except: return "-"

    def pct(num,den):
        try:
            n,d=float(num),float(den)
            if pd.isna(n) or pd.isna(d) or d==0: return "-"
            return f"{n/d*100:.1f}%"
        except: return "-"

    def qty_fmt(x):
        try: return f"{int(float(x)):,}" if float(x)!=0 else "-"
        except: return "-"

    def asp_fmt(x):
        try:
            f=float(x)
            if pd.isna(f) or f==0: return "-"
            return f"₹{f:,.0f}"
        except: return "-"

    def bracket(x):
        try:
            f=float(x)/1e5
            if pd.isna(f): return "-"
            if f==0: return "-"
            return f"({abs(f):,.2f})" if f<0 else f"{f:,.2f}"
        except: return "-"

    # ── Build the sheet rows ──────────────────────────────────────────────────
    all_rows = []

    # Title
    all_rows.append(["KENAZ PERFUMES — P&L DASHBOARD (INR Lacs)"] + [""]*( len(months)+1))
    all_rows.append(["Last updated: " + date.today().strftime("%d %b %Y")] + [""]*( len(months)+1))
    all_rows.append([""])  # blank

    # Header row
    all_rows.append(["Particulars"] + months + ["Total"])

    def row(label, vals, total, note=""):
        return [label] + [str(v) for v in vals] + [str(total)]

    def gap(): return [""] * (len(months)+2)

    # ── Data rows ─────────────────────────────────────────────────────────────
    all_rows.append(gap())
    all_rows.append(row("MRP Sales (₹ Lacs)",
        [lacs(v("MRP Sales",m)) for m in months],
        lacs(sum(v("MRP Sales",m) for m in months))))

    all_rows.append(row("Quantity (Units)",
        [qty_fmt(v("Quantity",m)) for m in months],
        qty_fmt(tot_qty)))

    all_rows.append(row("ASP (₹)",
        [asp_fmt(safe_div(v("Net Sales",m),v("Quantity",m))) for m in months],
        asp_fmt(safe_div(sum(v("Net Sales",m) for m in months), tot_qty))))

    all_rows.append(row("Discount %",
        [pct(v("MRP Sales",m)-v("Net Sales",m), v("MRP Sales",m)) for m in months],
        pct(sum(v("MRP Sales",m)-v("Net Sales",m) for m in months), sum(v("MRP Sales",m) for m in months))))

    all_rows.append(gap())
    all_rows.append(row("NET SALES (₹ Lacs)",
        [lacs(v("Net Sales",m)) for m in months],
        lacs(tot_ns)))

    all_rows.append(gap())
    all_rows.append(row("Less: COGS",
        [lacs(v("COGS",m)) for m in months],
        lacs(sum(v("COGS",m) for m in months))))
    all_rows.append(row("  COGS %",
        [pct(v("COGS",m),nsv(m)) for m in months],
        pct(sum(v("COGS",m) for m in months),tot_ns)))

    all_rows.append(gap())
    all_rows.append(row("MATERIAL MARGINS",
        [lacs(mat(m)) for m in months],
        lacs(sum(mat(m) for m in months))))
    all_rows.append(row("  Material Margins %",
        [pct(mat(m),nsv(m)) for m in months],
        pct(sum(mat(m) for m in months),tot_ns)))

    all_rows.append(gap())
    all_rows.append(row("Less: Freight Inwards",
        [lacs(v("Freight Inward",m)) for m in months],
        lacs(sum(v("Freight Inward",m) for m in months))))
    all_rows.append(row("Less: Wages - Fixed",
        [lacs(v("Wages",m)) for m in months],
        lacs(sum(v("Wages",m) for m in months))))
    all_rows.append(row("  Freight & Wages Total",
        [lacs(fw(m)) for m in months],
        lacs(sum(fw(m) for m in months))))
    all_rows.append(row("  Inward %age",
        [pct(fw(m),nsv(m)) for m in months],
        pct(sum(fw(m) for m in months),tot_ns)))

    all_rows.append(gap())
    all_rows.append(row("GROSS MARGINS",
        [lacs(gm(m)) for m in months],
        lacs(sum(gm(m) for m in months))))
    all_rows.append(row("  Gross Margins %",
        [pct(gm(m),nsv(m)) for m in months],
        pct(sum(gm(m) for m in months),tot_ns)))

    all_rows.append(gap())
    all_rows.append(row("Less: Commission Expense",
        [lacs(v("Commission",m)) for m in months],
        lacs(sum(v("Commission",m) for m in months))))
    all_rows.append(row("Less: Payment Gateway",
        [lacs(v("Payment Gateway",m)) for m in months],
        lacs(sum(v("Payment Gateway",m) for m in months))))
    all_rows.append(row("Less: Shipping Charges",
        [lacs(v("Shipping",m)) for m in months],
        lacs(sum(v("Shipping",m) for m in months))))
    all_rows.append(row("Less: Bulk Logistic",
        [lacs(v("Bulk Logistic",m)) for m in months],
        lacs(sum(v("Bulk Logistic",m) for m in months))))
    all_rows.append(row("Less: Packaging Cost",
        [lacs(v("Packaging",m)) for m in months],
        lacs(sum(v("Packaging",m) for m in months))))
    all_rows.append(row("Less: Warehousing",
        [lacs(v("Warehousing",m)) for m in months],
        lacs(sum(v("Warehousing",m) for m in months))))
    all_rows.append(row("Less: Rebate",
        [lacs(v("Rebate",m)) for m in months],
        lacs(sum(v("Rebate",m) for m in months))))
    all_rows.append(row("Less: Others",
        [lacs(v("Others",m)) for m in months],
        lacs(sum(v("Others",m) for m in months))))
    all_rows.append(row("  C&L Total",
        [lacs(cnl(m)) for m in months],
        lacs(sum(cnl(m) for m in months))))
    all_rows.append(row("  Commission & Logistics %",
        [pct(cnl(m),nsv(m)) for m in months],
        pct(sum(cnl(m) for m in months),tot_ns)))

    all_rows.append(gap())
    all_rows.append(row("CM1",
        [bracket(cm1(m)*1e5) for m in months],
        bracket(sum(cm1(m) for m in months)*1e5)))
    all_rows.append(row("  CM1 %",
        [pct(cm1(m),nsv(m)) for m in months],
        pct(sum(cm1(m) for m in months),tot_ns)))

    all_rows.append(gap())
    all_rows.append(row("Less: Performance Marketing (Ad Spend)",
        [lacs(v("Ad Spend",m)) for m in months],
        lacs(sum(v("Ad Spend",m) for m in months))))
    all_rows.append(row("  ACOS %",
        [pct(v("Ad Spend",m),nsv(m)) for m in months],
        pct(sum(v("Ad Spend",m) for m in months),tot_ns)))

    all_rows.append(gap())
    all_rows.append(row("CM2",
        [bracket(cm2(m)*1e5) for m in months],
        bracket(sum(cm2(m) for m in months)*1e5)))
    all_rows.append(row("  CM2 %",
        [pct(cm2(m),nsv(m)) for m in months],
        pct(sum(cm2(m) for m in months),tot_ns)))

    if exp_df is not None and not exp_df.empty:
        all_rows.append(gap())
        all_rows.append(row("Less: Brand Marketing (Influencer + Marketing)",
            [lacs(brand_mkt(m)) for m in months],
            lacs(sum(brand_mkt(m) for m in months))))
        all_rows.append(gap())
        all_rows.append(row("CM3",
            [bracket(cm3(m)*1e5) for m in months],
            bracket(sum(cm3(m) for m in months)*1e5)))
        all_rows.append(row("  CM3 %",
            [pct(cm3(m),nsv(m)) for m in months],
            pct(sum(cm3(m) for m in months),tot_ns)))

    all_rows.append(gap())
    all_rows.append(["Note: All values in INR Lacs. Negative values shown in (brackets). % = % of Net Sales."])

    # Write to sheet
    ws.update(all_rows)

    # ── Apply formatting ──────────────────────────────────────────────────────
    import gspread.utils as gu

    n_cols = len(months) + 2  # Particulars + months + Total

    def col_letter(n):
        return gu.rowcol_to_a1(1, n)[:-1]

    def fmt_req(range_str, bold=False, bg=None, fg=None, size=10,
                halign="LEFT", valign="MIDDLE", italic=False):
        fmt = {
            "horizontalAlignment": halign,
            "verticalAlignment":   valign,
            "textFormat": {
                "bold":     bold,
                "italic":   italic,
                "fontSize": size,
                "foregroundColor": _hex_to_rgb(fg or "FFFFFF"),
            },
        }
        if bg:
            fmt["backgroundColor"] = _hex_to_rgb(bg)
        return {"range": range_str, "format": fmt}

    def _hex_to_rgb(h):
        h = h.lstrip("#")
        if len(h) == 6:
            r,g,b = int(h[0:2],16)/255, int(h[2:4],16)/255, int(h[4:6],16)/255
            return {"red":r,"green":g,"blue":b}
        return {"red":1,"green":1,"blue":1}

    formats = []

    # Title row (row 1)
    formats.append(fmt_req(f"A1:{col_letter(n_cols)}1",
                           bold=True, bg="C9A84C", fg="1A1A1A", size=12, halign="LEFT"))

    # Updated row (row 2)
    formats.append(fmt_req(f"A2:{col_letter(n_cols)}2",
                           italic=True, bg="1A1A1A", fg="888888", size=9))

    # Header row (row 4)
    formats.append(fmt_req(f"A4:{col_letter(n_cols)}4",
                           bold=True, bg="C9A84C", fg="1A1A1A", size=10, halign="CENTER"))
    formats.append(fmt_req("A4", bold=True, bg="C9A84C", fg="1A1A1A", size=10, halign="LEFT"))

    # Whole sheet background
    formats.append(fmt_req(f"A1:{col_letter(n_cols)}200",
                           bg="1A1A1A", fg="CCCCCC", size=10, halign="RIGHT"))
    formats.append(fmt_req(f"A1:{col_letter(n_cols)}200",
                           bg="1A1A1A", fg="CCCCCC", size=10, halign="LEFT"))  # will override col A

    # Highlight total rows — NET SALES, MATERIAL MARGINS, GROSS MARGINS, CM1, CM2, CM3
    highlight_labels = ["NET SALES (₹ Lacs)","MATERIAL MARGINS","GROSS MARGINS","CM1","CM2","CM3"]
    for ri, r in enumerate(all_rows, 1):
        if r and r[0] in highlight_labels:
            formats.append(fmt_req(f"A{ri}:{col_letter(n_cols)}{ri}",
                                   bold=True, bg="2A2A2A", fg="C9A84C", size=10, halign="RIGHT"))
            formats.append(fmt_req(f"A{ri}", bold=True, bg="2A2A2A", fg="C9A84C", size=10, halign="LEFT"))

        # Percent rows (indented with spaces)
        if r and str(r[0]).startswith("  "):
            formats.append(fmt_req(f"A{ri}:{col_letter(n_cols)}{ri}",
                                   italic=True, bg="111111", fg="999999", size=9, halign="RIGHT"))
            formats.append(fmt_req(f"A{ri}", italic=True, bg="111111", fg="999999", size=9, halign="LEFT"))

    # Total column — last col
    last_col = col_letter(n_cols)
    formats.append(fmt_req(f"{last_col}4:{last_col}200",
                           bold=True, bg="1E1E1E", fg="C9A84C", halign="RIGHT"))

    # Apply all formats in one batch call
    ws.batch_format(formats)

    # Column widths
    requests = [
        {"updateDimensionProperties": {
            "range": {"sheetId": ws.id, "dimension": "COLUMNS", "startIndex": 0, "endIndex": 1},
            "properties": {"pixelSize": 280},
            "fields": "pixelSize"
        }},
    ]
    for ci in range(1, n_cols):
        requests.append({"updateDimensionProperties": {
            "range": {"sheetId": ws.id, "dimension": "COLUMNS",
                      "startIndex": ci, "endIndex": ci+1},
            "properties": {"pixelSize": 100},
            "fields": "pixelSize"
        }})
    # Freeze row 4 (header)
    requests.append({"updateSheetProperties": {
        "properties": {"sheetId": ws.id, "gridProperties": {"frozenRowCount": 4, "frozenColumnCount": 1}},
        "fields": "gridProperties.frozenRowCount,gridProperties.frozenColumnCount"
    }})
    sh.batch_update({"requests": requests})

    return ws.url if hasattr(ws, 'url') else f"https://docs.google.com/spreadsheets/d/{SHEET_KEY}"

def save_expenses_to_gsheet(client, exp_df: pd.DataFrame):
    sh=get_sheet(client)
    try: ws=sh.worksheet("Expenses")
    except gspread.exceptions.WorksheetNotFound:
        ws=sh.add_worksheet(title="Expenses",rows=1000,cols=10)
    existing=ws.get_all_records()
    exp_df=exp_df.copy()
    def df_to_rows(df): return [df.columns.tolist()]+df.astype(str).values.tolist()
    if not existing:
        ws.update(df_to_rows(exp_df))
        return len(exp_df),0
    ex=pd.DataFrame(existing)
    ex_keys=ex["Month"].astype(str)+"_"+ex["Vendor"].astype(str)+"_"+ex["Section"].astype(str)
    new_keys=exp_df["Month"].astype(str)+"_"+exp_df["Vendor"].astype(str)+"_"+exp_df["Section"].astype(str)
    truly_new=exp_df[~new_keys.isin(ex_keys)]
    if len(truly_new)==0: return 0,len(exp_df)
    all_cols=list(dict.fromkeys(ex.columns.tolist()+truly_new.columns.tolist()))
    combined=pd.concat([ex.reindex(columns=all_cols,fill_value=""),
                        truly_new.reindex(columns=all_cols,fill_value="")],ignore_index=True)
    combined=combined.sort_values(["Month","Section","Vendor"])
    ws.clear()
    ws.update(df_to_rows(combined))
    return len(truly_new),len(exp_df)-len(truly_new)

# ─── XLSB Parser ──────────────────────────────────────────────────────────────
def parse_xlsb(file_bytes: bytes) -> pd.DataFrame:
    with tempfile.NamedTemporaryFile(suffix=".xlsb",delete=False) as tmp:
        tmp.write(file_bytes); tmp_path=tmp.name
    try:
        rows=[]
        with open_workbook(tmp_path) as wb:
            with wb.get_sheet("For P&L") as ws:
                for row in ws.rows(): rows.append([c.v for c in row])
    finally: os.unlink(tmp_path)
    headers=rows[4]
    df=pd.DataFrame(rows[5:],columns=headers)
    df=df[pd.to_numeric(df["Sum of Revenue Without Tax"],errors="coerce").notna()]
    df=df[pd.to_numeric(df["Sum of Revenue Without Tax"],errors="coerce")!=0]
    df=df.dropna(subset=["Channel"])
    df=df[df["Channel"].astype(str).isin(["Website","FBA","RK","Meesho","Flipkart","Myntra PPMP"])]
    for col in [c for c in df.columns if isinstance(c,str) and c.startswith("Sum of")]:
        df[col]=pd.to_numeric(df[col],errors="coerce").fillna(0)
    df["Channel"]=df["Channel"].apply(lambda x:"Amazon" if x in ["FBA","RK"] else x)
    df["Month_date"]=df["Month"].apply(xlsb_to_date)
    df["Month_name"]=df["Month_date"].apply(lambda d:d.strftime("%b-%y") if d else "Unknown")
    df["Month_serial"]=df["Month"].apply(lambda x:str(int(float(x))) if pd.notna(x) else "")
    grp=df.groupby(["Month_serial","Month_name","Channel"])[
        ["Sum of TOTAL MRP","Sum of Qty","Sum of Revenue Without Tax","Sum of COGS",
         "Sum of Inward","Sum of Wages","Sum of commission","Sum of Payment Gateway",
         "Sum of Shipping","Sum of others","Sum of Total Spend","Sum of Bulk Logistic Cost",
         "Sum of Packaging Cost","Sum of Warehousing Charges","Sum of Rebate"]
    ].sum().reset_index()
    return pd.DataFrame({
        "Month_serial":grp["Month_serial"],"Month_name":grp["Month_name"],"Channel":grp["Channel"],
        "MRP Sales":grp["Sum of TOTAL MRP"],"Quantity":grp["Sum of Qty"],
        "Net Sales":grp["Sum of Revenue Without Tax"],"COGS":grp["Sum of COGS"],
        "Freight Inward":grp["Sum of Inward"],"Wages":grp["Sum of Wages"],
        "Commission":grp["Sum of commission"],"Payment Gateway":grp["Sum of Payment Gateway"],
        "Shipping":grp["Sum of Shipping"],"Others":grp["Sum of others"],
        "Ad Spend":grp["Sum of Total Spend"],"Bulk Logistic":grp["Sum of Bulk Logistic Cost"],
        "Packaging":grp["Sum of Packaging Cost"],"Warehousing":grp["Sum of Warehousing Charges"],
        "Rebate":grp["Sum of Rebate"],
    })

def parse_sku_data(file_bytes: bytes) -> pd.DataFrame:
    with tempfile.NamedTemporaryFile(suffix=".xlsb",delete=False) as tmp:
        tmp.write(file_bytes); tmp_path=tmp.name
    try:
        rows=[]
        with open_workbook(tmp_path) as wb:
            with wb.get_sheet("Data") as ws:
                for row in ws.rows(): rows.append([c.v for c in row])
    finally: os.unlink(tmp_path)
    headers=rows[1]
    df=pd.DataFrame(rows[2:],columns=headers)
    df["Month_label"]=df["Month"].apply(
        lambda n:(date(1899,12,30)+timedelta(days=int(float(n)))).strftime("%b-%y") if pd.notna(n) else None)
    df["EAN"]=pd.to_numeric(df["New SKU"],errors="coerce")
    df["Model"]=df["EAN"].apply(lambda x:EAN_MAP.get(int(x),"Other") if pd.notna(x) else "Unknown")
    df["Channel"]=df["Channel"].apply(lambda x:"Amazon" if x in ["FBA","RK"] else str(x))
    for c in SKU_NUM_COLS:
        if c in df.columns: df[c]=pd.to_numeric(df[c],errors="coerce").fillna(0)
    df=df.dropna(subset=["Month_label"])
    sales_df=df[df["Sale / Return"]=="Sale"]
    returns_df=df[df["Sale / Return"]=="Return"]
    grp=sales_df.groupby(["Model","EAN","Month_label","Month","Channel"])[SKU_NUM_COLS].sum().reset_index()
    ret_grp=returns_df.groupby(["Model","EAN","Month_label","Month","Channel"])[
        ["Revenue Without Tax"]].sum().reset_index().rename(columns={"Revenue Without Tax":"Return Amount"})
    grp=grp.merge(ret_grp,on=["Model","EAN","Month_label","Month","Channel"],how="left")
    grp["Return Amount"]=grp["Return Amount"].fillna(0)
    grp["Month_sort"]=grp["Month"].apply(lambda s:str(int(float(s))).zfill(10) if pd.notna(s) else "9999999999")
    grp=grp.rename(columns={
        "Revenue Without Tax":"Net Sales","Qty":"Quantity","TOTAL MRP":"MRP Sales",
        "Inward":"Freight Inward","commission":"Commission","Bulk Logistic Cost":"Bulk Logistic",
        "Packaging Cost":"Packaging","Warehousing Charges":"Warehousing","others":"Others","Total Spend":"Ad Spend"})
    grp["RTO%"]=(grp["Return Amount"]/grp["Net Sales"].replace(0,np.nan)*100).fillna(0)
    return grp

def enrich(df: pd.DataFrame) -> pd.DataFrame:
    df=df.copy()
    for col in ["Net Sales","COGS","Freight Inward","Wages","Commission","Payment Gateway",
                "Shipping","Others","Ad Spend","Bulk Logistic","Packaging","Warehousing","Rebate","MRP Sales","Quantity"]:
        if col in df.columns: df[col]=pd.to_numeric(df[col],errors="coerce").fillna(0)
    df["Material Margin"]=df["Net Sales"]-df["COGS"]
    df["Freight+Wages"]=df["Freight Inward"]+df["Wages"]
    df["GM"]=df["Material Margin"]-df["Freight+Wages"]
    df["CnL"]=(df["Commission"]+df["Payment Gateway"]+df["Shipping"]+
               df["Bulk Logistic"]+df["Packaging"]+df["Warehousing"]+df["Rebate"]+df["Others"])
    df["CM1"]=df["GM"]-df["CnL"]
    df["CM2"]=df["CM1"]-df["Ad Spend"]
    safe=df["Net Sales"].replace(0,np.nan)
    df["Discount%"]=(1-df["Net Sales"]/df["MRP Sales"].replace(0,np.nan))*100
    df["COGS%"]=df["COGS"]/safe*100
    df["GM%"]=df["GM"]/safe*100
    df["CM1%"]=df["CM1"]/safe*100
    df["ACOS%"]=df["Ad Spend"]/safe*100
    df["CM2%"]=df["CM2"]/safe*100
    df["ASP"]=df["Net Sales"]/df["Quantity"].replace(0,np.nan)
    df["Month_sort"]=df["Month_serial"].apply(
        lambda s:str(int(float(s))).zfill(10) if str(s).strip() not in ["","nan"] else "9999999999")
    return df

# ─── P&L Table Builder ────────────────────────────────────────────────────────
def build_pnl_table(df,months,channels):
    METRICS=["MRP Sales","Quantity","Net Sales","COGS","Freight Inward","Wages",
             "Commission","Payment Gateway","Shipping","Bulk Logistic","Packaging",
             "Warehousing","Rebate","Others","Ad Spend"]
    grp=df.groupby(["Month_name","Month_sort"])[METRICS].sum().reset_index()
    grp=grp[grp["Month_name"].isin(months)].copy()
    month_order=grp.sort_values("Month_sort")["Month_name"].tolist()
    lookup={}
    for _,row in grp.iterrows():
        lookup[row["Month_name"]]={m:float(row[m]) for m in METRICS}
    def v(metric,m): return lookup.get(m,{}).get(metric,0)
    def nsv(m): val=v("Net Sales",m); return val if val!=0 else np.nan
    def mrp(m): val=v("MRP Sales",m); return val if val!=0 else np.nan
    def mat_margin(m): return v("Net Sales",m)-v("COGS",m)
    def fw(m): return v("Freight Inward",m)+v("Wages",m)
    def gm(m): return mat_margin(m)-fw(m)
    def cnl(m): return (v("Commission",m)+v("Payment Gateway",m)+v("Shipping",m)+
                        v("Bulk Logistic",m)+v("Packaging",m)+v("Warehousing",m)+
                        v("Rebate",m)+v("Others",m))
    def cm1(m): return gm(m)-cnl(m)
    def cm2(m): return cm1(m)-v("Ad Spend",m)
    tot={metric:sum(lookup.get(m,{}).get(metric,0) for m in month_order) for metric in METRICS}
    def T(fn): return sum(fn(m) for m in month_order)
    tot_ns=tot.get("Net Sales",0); tot_mrp=tot.get("MRP Sales",0)

    def th(label):
        heads="".join(f"<th>{m}</th>" for m in month_order)
        return f"<thead><tr><th>{label}</th>{heads}<th>Total</th></tr></thead>"

    def data_row(label,vals_fn,total_fn,fmt_fn=L,cls="",color_fn=None,inverse=False):
        cells=""
        for m in month_order:
            val=vals_fn(m); s=fmt_fn(val)
            if color_fn: c=color_val(val,inverse); cells+=f"<td style='color:{c}'>{s}</td>"
            else: cells+=f"<td>{s}</td>"
        tv=total_fn(); ts=fmt_fn(tv)
        if color_fn: tc=color_val(tv,inverse); return f"<tr class='{cls}'><td>{label}</td>{cells}<td style='color:{tc}'>{ts}</td></tr>"
        return f"<tr class='{cls}'><td>{label}</td>{cells}<td>{ts}</td></tr>"

    def pct_row(label,pct_fn,total_pct_fn,cls="pct-row",inverse=False):
        cells=""
        for m in month_order:
            val=pct_fn(m); c=color_val(val,inverse); cells+=f"<td style='color:{c}'>{P(val)}</td>"
        tv=total_pct_fn(); tc=color_val(tv,inverse)
        return f"<tr class='{cls}'><td>{label}</td>{cells}<td style='color:{tc}'>{P(tv)}</td></tr>"

    def gap():
        n=len(month_order)+2
        return f"<tr class='section-gap'>{'<td></td>'*n}</tr>"

    rows_html=""
    rows_html+=data_row("MRP Sales",lambda m:v("MRP Sales",m)/1e5,lambda:tot_mrp/1e5)
    rows_html+=data_row("Quantity",lambda m:v("Quantity",m),lambda:tot.get("Quantity",0),
                        fmt_fn=lambda x:f"{int(x):,}" if not pd.isna(x) and x!=0 else "-")
    rows_html+=data_row("ASP",lambda m:safe_div(v("Net Sales",m),v("Quantity",m)),
                        lambda:safe_div(tot_ns,tot.get("Quantity",0)),
                        fmt_fn=lambda x:f"&#8377;{x:,.0f}" if not pd.isna(x) else "-")
    rows_html+=pct_row("Discount %",
                       lambda m:(1-safe_div(v("Net Sales",m),v("MRP Sales",m)))*100 if v("MRP Sales",m) else np.nan,
                       lambda:(1-safe_div(tot_ns,tot_mrp))*100 if tot_mrp else np.nan,inverse=True)
    rows_html+=gap()
    rows_html+=data_row("Net Sales",lambda m:v("Net Sales",m)/1e5,lambda:tot_ns/1e5,cls="total-row")
    rows_html+=gap()
    rows_html+=data_row("Less: COGS",lambda m:v("COGS",m)/1e5,lambda:tot.get("COGS",0)/1e5,color_fn=True,inverse=True)
    rows_html+=pct_row("COGS %",lambda m:safe_div(v("COGS",m),nsv(m))*100,
                       lambda:safe_div(tot.get("COGS",0),tot_ns)*100 if tot_ns else np.nan,inverse=True)
    rows_html+=gap()
    rows_html+=data_row("Material Margins",lambda m:mat_margin(m)/1e5,lambda:T(mat_margin)/1e5,cls="total-row",color_fn=True)
    rows_html+=pct_row("Material Margins (%)",lambda m:safe_div(mat_margin(m),nsv(m))*100,
                       lambda:safe_div(T(mat_margin),tot_ns)*100 if tot_ns else np.nan)
    rows_html+=gap()
    rows_html+=data_row("Less: Freight Inwards",lambda m:v("Freight Inward",m)/1e5,lambda:tot.get("Freight Inward",0)/1e5,color_fn=True,inverse=True)
    rows_html+=data_row("Less: Wages - Fixed",lambda m:v("Wages",m)/1e5,lambda:tot.get("Wages",0)/1e5,color_fn=True,inverse=True)
    rows_html+=data_row("Freight Inwards & Wages Total",lambda m:fw(m)/1e5,lambda:T(fw)/1e5,color_fn=True,inverse=True)
    rows_html+=pct_row("Inward %age",lambda m:safe_div(fw(m),nsv(m))*100,
                       lambda:safe_div(T(fw),tot_ns)*100 if tot_ns else np.nan,inverse=True)
    rows_html+=gap()
    rows_html+=data_row("Gross Margins",lambda m:gm(m)/1e5,lambda:T(gm)/1e5,cls="total-row",color_fn=True)
    rows_html+=pct_row("Gross Margins (%)",lambda m:safe_div(gm(m),nsv(m))*100,
                       lambda:safe_div(T(gm),tot_ns)*100 if tot_ns else np.nan)
    rows_html+=gap()
    rows_html+=data_row("Less: Commission Expense",lambda m:v("Commission",m)/1e5,lambda:tot.get("Commission",0)/1e5,color_fn=True,inverse=True)
    rows_html+=data_row("Less: Payment Gateway Charges",lambda m:v("Payment Gateway",m)/1e5,lambda:tot.get("Payment Gateway",0)/1e5,color_fn=True,inverse=True)
    rows_html+=data_row("Less: Shipping Charges",lambda m:v("Shipping",m)/1e5,lambda:tot.get("Shipping",0)/1e5,color_fn=True,inverse=True)
    rows_html+=data_row("Less: Bulk Logistic Cost",lambda m:v("Bulk Logistic",m)/1e5,lambda:tot.get("Bulk Logistic",0)/1e5,color_fn=True,inverse=True)
    rows_html+=data_row("Less: Packaging Cost",lambda m:v("Packaging",m)/1e5,lambda:tot.get("Packaging",0)/1e5,color_fn=True,inverse=True)
    rows_html+=data_row("Less: Warehousing Charges",lambda m:v("Warehousing",m)/1e5,lambda:tot.get("Warehousing",0)/1e5,color_fn=True,inverse=True)
    rows_html+=data_row("Less: Rebate",lambda m:v("Rebate",m)/1e5,lambda:tot.get("Rebate",0)/1e5,color_fn=True,inverse=True)
    rows_html+=data_row("Less: Others",lambda m:v("Others",m)/1e5,lambda:tot.get("Others",0)/1e5,color_fn=True,inverse=True)
    rows_html+=data_row("Commission & Logistics Total",lambda m:cnl(m)/1e5,lambda:T(cnl)/1e5,color_fn=True,inverse=True)
    rows_html+=pct_row("Commission & Logistics %",lambda m:safe_div(cnl(m),nsv(m))*100,
                       lambda:safe_div(T(cnl),tot_ns)*100 if tot_ns else np.nan,inverse=True)
    rows_html+=gap()
    rows_html+=data_row("CM1",lambda m:cm1(m)/1e5,lambda:T(cm1)/1e5,cls="total-row",color_fn=True,
                        fmt_fn=lambda x:Lbold(x) if not pd.isna(x) else "-")
    rows_html+=pct_row("CM1 (%)",lambda m:safe_div(cm1(m),nsv(m))*100,
                       lambda:safe_div(T(cm1),tot_ns)*100 if tot_ns else np.nan)
    rows_html+=gap()
    rows_html+=data_row("Less: Performance Marketing",lambda m:v("Ad Spend",m)/1e5,lambda:tot.get("Ad Spend",0)/1e5,color_fn=True,inverse=True)
    rows_html+=pct_row("ACOS (%)",lambda m:safe_div(v("Ad Spend",m),nsv(m))*100,
                       lambda:safe_div(tot.get("Ad Spend",0),tot_ns)*100 if tot_ns else np.nan,inverse=True)
    rows_html+=gap()
    rows_html+=data_row("CM2",lambda m:cm2(m)/1e5,lambda:T(cm2)/1e5,cls="total-row",color_fn=True,
                        fmt_fn=lambda x:Lbold(x) if not pd.isna(x) else "-")
    rows_html+=pct_row("CM2 (%)",lambda m:safe_div(cm2(m),nsv(m))*100,
                       lambda:safe_div(T(cm2),tot_ns)*100 if tot_ns else np.nan)

    exp_df=st.session_state.get("expense_df",pd.DataFrame())
    if not exp_df.empty:
        rows_html+=gap()
        def brand_mkt(m):
            sub=exp_df[exp_df["Month"]==m]
            return float(sub["Amount"].sum()) if not sub.empty else 0.0
        def cm3(m): return cm2(m)-brand_mkt(m)
        def T_brand(): return sum(brand_mkt(m) for m in month_order)
        rows_html+=data_row("Less: Brand Marketing",lambda m:brand_mkt(m)/1e5,lambda:T_brand()/1e5,color_fn=True,inverse=True)
        rows_html+=data_row("CM3",lambda m:cm3(m)/1e5,lambda:T(cm3)/1e5,cls="total-row",color_fn=True,
                            fmt_fn=lambda x:Lbold(x) if not pd.isna(x) else "-")
        rows_html+=pct_row("CM3 (%)",lambda m:safe_div(cm3(m),nsv(m))*100,
                           lambda:safe_div(T(cm3),tot_ns)*100 if tot_ns else np.nan)

    return f"""<div style='overflow-x:auto'><table class='pnl-table'>{th("Particulars (INR Lacs)")}<tbody>{rows_html}</tbody></table></div>"""

# ─── Excel Export ─────────────────────────────────────────────────────────────
def build_pnl_excel(df,months):
    import openpyxl
    from openpyxl.styles import PatternFill,Font,Alignment
    from openpyxl.utils import get_column_letter
    METRICS=["MRP Sales","Quantity","Net Sales","COGS","Freight Inward","Wages",
             "Commission","Payment Gateway","Shipping","Bulk Logistic","Packaging",
             "Warehousing","Rebate","Others","Ad Spend"]
    grp=df.groupby(["Month_name","Month_sort"])[METRICS].sum().reset_index()
    grp=grp[grp["Month_name"].isin(months)].copy()
    lookup={row["Month_name"]:{m:float(row[m]) for m in METRICS} for _,row in grp.iterrows()}
    month_order=grp.sort_values("Month_sort")["Month_name"].tolist()
    def v(metric,m): return lookup.get(m,{}).get(metric,0)
    def nsv(m): return v("Net Sales",m) or np.nan
    def mat(m): return v("Net Sales",m)-v("COGS",m)
    def fw(m): return v("Freight Inward",m)+v("Wages",m)
    def gm(m): return mat(m)-fw(m)
    def cnl(m): return sum(v(c,m) for c in ["Commission","Payment Gateway","Shipping","Bulk Logistic","Packaging","Warehousing","Rebate","Others"])
    def cm1(m): return gm(m)-cnl(m)
    def cm2(m): return cm1(m)-v("Ad Spend",m)
    def tot_val(fn): return sum(fn(m) for m in month_order)
    tot_ns=sum(v("Net Sales",m) for m in month_order) or np.nan
    tot_qty=sum(v("Quantity",m) for m in month_order)
    GOLD_HEX="C9A84C"; DARK_HEX="1A1A1A"
    gold_fill=PatternFill("solid",fgColor=GOLD_HEX)
    dark_fill=PatternFill("solid",fgColor="1E1E1E")
    pct_fill=PatternFill("solid",fgColor="222222")
    tot_fill=PatternFill("solid",fgColor="2A2A2A")
    GREEN="4CAF50"; RED="EF5350"
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="P&L Summary"
    ws.sheet_view.showGridLines=False; ws.freeze_panes="B3"
    months_all=month_order+["Total"]; headers=["Particulars (INR Lacs)"]+months_all; ncols=len(headers)
    for ci,h in enumerate(headers,1):
        cell=ws.cell(1,ci,h); cell.fill=gold_fill
        cell.font=Font(bold=True,color=DARK_HEX,size=10)
        cell.alignment=Alignment(horizontal="center" if ci>1 else "left",vertical="center")
    ws.row_dimensions[1].height=22
    exp_df_ss=st.session_state.get("expense_df",pd.DataFrame())
    def brand_mkt_xl(m):
        if exp_df_ss.empty: return 0
        sub=exp_df_ss[exp_df_ss["Month"]==m]; return float(sub["Amount"].sum()) if not sub.empty else 0.0
    def cm3_xl(m): return cm2(m)-brand_mkt_xl(m)
    rows_def=[
        ("MRP Sales",[v("MRP Sales",m)/1e5 for m in month_order],sum(v("MRP Sales",m) for m in month_order)/1e5,"0.00",False,False,False),
        ("Quantity",[int(v("Quantity",m)) for m in month_order],int(tot_qty),"#,##0",False,False,False),
        ("ASP",[safe_div(v("Net Sales",m),v("Quantity",m)) for m in month_order],safe_div(sum(v("Net Sales",m) for m in month_order),tot_qty),"#,##0",False,False,False),
        ("Discount %",[(1-safe_div(v("Net Sales",m),v("MRP Sales",m)))*100 if v("MRP Sales",m) else np.nan for m in month_order],(1-safe_div(sum(v("Net Sales",m) for m in month_order),sum(v("MRP Sales",m) for m in month_order)))*100,"0.0%",False,True,True),
        (None,None,None,None,False,False,False),
        ("Net Sales",[v("Net Sales",m)/1e5 for m in month_order],tot_ns/1e5 if not pd.isna(tot_ns) else 0,"0.00",True,False,False),
        (None,None,None,None,False,False,False),
        ("Less: COGS",[v("COGS",m)/1e5 for m in month_order],sum(v("COGS",m) for m in month_order)/1e5,"0.00",False,False,True),
        ("COGS %",[safe_div(v("COGS",m),nsv(m))*100 for m in month_order],safe_div(sum(v("COGS",m) for m in month_order),tot_ns)*100,"0.0%",False,True,True),
        (None,None,None,None,False,False,False),
        ("Material Margins",[mat(m)/1e5 for m in month_order],tot_val(mat)/1e5,"0.00",True,False,False),
        ("Material Margins (%)",[safe_div(mat(m),nsv(m))*100 for m in month_order],safe_div(tot_val(mat),tot_ns)*100,"0.0%",False,True,False),
        (None,None,None,None,False,False,False),
        ("Less: Freight Inwards",[v("Freight Inward",m)/1e5 for m in month_order],sum(v("Freight Inward",m) for m in month_order)/1e5,"0.00",False,False,True),
        ("Less: Wages - Fixed",[v("Wages",m)/1e5 for m in month_order],sum(v("Wages",m) for m in month_order)/1e5,"0.00",False,False,True),
        ("Freight & Wages Total",[fw(m)/1e5 for m in month_order],tot_val(fw)/1e5,"0.00",False,False,True),
        ("Inward %age",[safe_div(fw(m),nsv(m))*100 for m in month_order],safe_div(tot_val(fw),tot_ns)*100,"0.0%",False,True,True),
        (None,None,None,None,False,False,False),
        ("Gross Margins",[gm(m)/1e5 for m in month_order],tot_val(gm)/1e5,"0.00",True,False,False),
        ("Gross Margins (%)",[safe_div(gm(m),nsv(m))*100 for m in month_order],safe_div(tot_val(gm),tot_ns)*100,"0.0%",False,True,False),
        (None,None,None,None,False,False,False),
        ("Less: Commission Expense",[v("Commission",m)/1e5 for m in month_order],sum(v("Commission",m) for m in month_order)/1e5,"0.00",False,False,True),
        ("Less: Payment Gateway",[v("Payment Gateway",m)/1e5 for m in month_order],sum(v("Payment Gateway",m) for m in month_order)/1e5,"0.00",False,False,True),
        ("Less: Shipping Charges",[v("Shipping",m)/1e5 for m in month_order],sum(v("Shipping",m) for m in month_order)/1e5,"0.00",False,False,True),
        ("Less: Bulk Logistic",[v("Bulk Logistic",m)/1e5 for m in month_order],sum(v("Bulk Logistic",m) for m in month_order)/1e5,"0.00",False,False,True),
        ("Less: Packaging Cost",[v("Packaging",m)/1e5 for m in month_order],sum(v("Packaging",m) for m in month_order)/1e5,"0.00",False,False,True),
        ("Less: Warehousing",[v("Warehousing",m)/1e5 for m in month_order],sum(v("Warehousing",m) for m in month_order)/1e5,"0.00",False,False,True),
        ("Less: Rebate",[v("Rebate",m)/1e5 for m in month_order],sum(v("Rebate",m) for m in month_order)/1e5,"0.00",False,False,True),
        ("Less: Others",[v("Others",m)/1e5 for m in month_order],sum(v("Others",m) for m in month_order)/1e5,"0.00",False,False,True),
        ("Commission & Logistics Total",[cnl(m)/1e5 for m in month_order],tot_val(cnl)/1e5,"0.00",False,False,True),
        ("Commission & Logistics %",[safe_div(cnl(m),nsv(m))*100 for m in month_order],safe_div(tot_val(cnl),tot_ns)*100,"0.0%",False,True,True),
        (None,None,None,None,False,False,False),
        ("CM1",[cm1(m)/1e5 for m in month_order],tot_val(cm1)/1e5,"0.00",True,False,False),
        ("CM1 (%)",[safe_div(cm1(m),nsv(m))*100 for m in month_order],safe_div(tot_val(cm1),tot_ns)*100,"0.0%",False,True,False),
        (None,None,None,None,False,False,False),
        ("Less: Performance Marketing",[v("Ad Spend",m)/1e5 for m in month_order],sum(v("Ad Spend",m) for m in month_order)/1e5,"0.00",False,False,True),
        ("ACOS (%)",[safe_div(v("Ad Spend",m),nsv(m))*100 for m in month_order],safe_div(sum(v("Ad Spend",m) for m in month_order),tot_ns)*100,"0.0%",False,True,True),
        (None,None,None,None,False,False,False),
        ("CM2",[cm2(m)/1e5 for m in month_order],tot_val(cm2)/1e5,"0.00",True,False,False),
        ("CM2 (%)",[safe_div(cm2(m),nsv(m))*100 for m in month_order],safe_div(tot_val(cm2),tot_ns)*100,"0.0%",False,True,False),
    ]
    if not exp_df_ss.empty:
        rows_def+=[(None,None,None,None,False,False,False),
                   ("Less: Brand Marketing",[brand_mkt_xl(m)/1e5 for m in month_order],sum(brand_mkt_xl(m) for m in month_order)/1e5,"0.00",False,False,True),
                   ("CM3",[cm3_xl(m)/1e5 for m in month_order],tot_val(cm3_xl)/1e5,"0.00",True,False,False),
                   ("CM3 (%)",[safe_div(cm3_xl(m),nsv(m))*100 for m in month_order],safe_div(tot_val(cm3_xl),tot_ns)*100,"0.0%",False,True,False)]
    for ri,row_def in enumerate(rows_def,2):
        label,vals,total,fmt,is_total,is_pct,inverse=row_def
        if label is None:
            ws.row_dimensions[ri].height=4
            for ci in range(1,ncols+1): ws.cell(ri,ci).fill=PatternFill("solid",fgColor="111111")
            continue
        lc=ws.cell(ri,1,label)
        if is_total: lc.fill=tot_fill; lc.font=Font(bold=True,color=GOLD_HEX,size=10)
        elif is_pct: lc.fill=pct_fill; lc.font=Font(color="888888",size=9,italic=True)
        else: lc.fill=dark_fill; lc.font=Font(color="CCCCCC",size=10)
        lc.alignment=Alignment(horizontal="left",vertical="center",indent=1)
        all_vals=list(vals)+[total]
        for ci,val in enumerate(all_vals,2):
            cell=ws.cell(ri,ci)
            try: fv=float(val) if val is not None and not(isinstance(val,float) and pd.isna(val)) else None
            except: fv=None
            if fv is not None: cell.value=fv/100 if is_pct else fv
            else: cell.value=None
            cell.number_format="0.0%" if is_pct else fmt
            if fv is not None and not pd.isna(fv):
                is_pos=fv>=0
                if inverse: is_pos=not is_pos
                font_color=GREEN if is_pos else RED
            else: font_color="CCCCCC"
            if is_total: cell.fill=tot_fill; cell.font=Font(bold=True,color=font_color,size=10)
            elif is_pct: cell.fill=pct_fill; cell.font=Font(color=font_color,size=9,italic=True)
            else: cell.fill=dark_fill; cell.font=Font(color=font_color,size=10)
            cell.alignment=Alignment(horizontal="right",vertical="center")
        ws.row_dimensions[ri].height=18
    ws.column_dimensions["A"].width=32
    for ci in range(2,ncols+2): ws.column_dimensions[get_column_letter(ci)].width=12
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

# ─── Expense Parser ───────────────────────────────────────────────────────────
def parse_expense_data(file_bytes: bytes) -> pd.DataFrame:
    import openpyxl as _xl
    wb=_xl.load_workbook(io.BytesIO(file_bytes),read_only=True)
    SHEET_MAP={"Sep - 25":"Sep-25","Oct - 25":"Oct-25","Nov - 25":"Nov-25",
               "Dec - 25":"Dec-25","Jan - 26":"Jan-26","Feb-26":"Feb-26",
               "March-26":"Mar-26","April-26":"Apr-26"}
    SKIP={"Name of the vendor","Website Spends","Marketplaces Spends","Marketplaces Spend","Total","","None"}
    records=[]
    for sheet_name,month_label in SHEET_MAP.items():
        if sheet_name not in wb.sheetnames: continue
        ws=wb[sheet_name]; current_section=None
        for row in ws.iter_rows(values_only=True):
            row=list(row)+[None]*20
            col_b=str(row[1]).strip() if row[1] else ""
            col_k=row[10]
            non_null=[vv for vv in row if vv is not None]
            if len(non_null)==1 and col_b: current_section=col_b.lower().strip(); continue
            if col_b in SKIP or not col_b or not current_section: continue
            if "influencer" not in current_section and "marketing" not in current_section: continue
            try: amt=float(col_k) if col_k and str(col_k).lower() not in ["none","pending",""] else 0.0
            except: amt=0.0
            records.append({"Month":month_label,
                            "Section":"Influencer Spend" if "influencer" in current_section else "Marketing Spend",
                            "Vendor":col_b,"Nature":str(row[3]).strip() if row[3] else "","Amount":amt})
    return pd.DataFrame(records)

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown(f"""<div style='text-align:center;padding:16px 0 8px 0'>
      <div style='font-size:26px'>🕌</div>
      <div style='font-size:17px;font-weight:800;color:{GOLD}'>Kenaz Perfumes</div>
      <div style='font-size:11px;color:#666;letter-spacing:1px;text-transform:uppercase;margin-top:3px'>One Guardian Brands</div>
    </div>""",unsafe_allow_html=True)
    st.markdown("---")
    st.markdown(f"📋 **Sheet:** [Kenaz_PL_DB](https://docs.google.com/spreadsheets/d/{SHEET_KEY})")

    st.markdown("### 📤 Upload P&L File")
    uploaded=st.file_uploader("Kenaz P&L (.xlsb)",type=["xlsb"])
    if uploaded:
        try:
            file_bytes=uploaded.read()
            st.session_state["last_upload_bytes"]=file_bytes
            raw=parse_xlsb(file_bytes)
            st.session_state["parsed_df"]=enrich(raw)
            st.session_state["sku_df_cache"]=parse_sku_data(file_bytes)
            ch_counts=raw["Channel"].value_counts()
            st.success(f"✅ {len(raw):,} rows | {raw['Month_name'].nunique()} months")
            st.info("  \n".join(f"• {ch}: {cnt}" for ch,cnt in ch_counts.items()))
            if st.button("💾 Save to Google Sheets",type="primary"):
                with st.spinner("Saving data + writing formatted P&L sheet…"):
                    client=get_gsheet_client()
                    added,dupes=save_to_gsheet(client,raw)
                    exp_df_save=st.session_state.get("expense_df",pd.DataFrame())
                    write_pnl_sheet(client,st.session_state["parsed_df"],
                                    exp_df_save if not exp_df_save.empty else None)
                st.success(f"✅ {added:,} new rows saved. Formatted P&L sheet updated.")
                st.cache_data.clear()
        except Exception as e:
            st.error(f"Parse error: {e}")

    st.markdown("### 📊 Upload Expense Sheet")
    uploaded_exp=st.file_uploader("Expense Sheet (.xlsx)",type=["xlsx"],key="exp_upload")
    if uploaded_exp:
        try:
            exp_bytes=uploaded_exp.read()
            parsed_exp=parse_expense_data(exp_bytes)
            st.session_state["expense_df"]=parsed_exp
            st.success(f"✅ {len(parsed_exp)} expense records loaded")
            if st.button("💾 Save Expenses to Google Sheets",type="primary",key="save_exp_btn"):
                with st.spinner("Saving expenses…"):
                    client=get_gsheet_client()
                    added,dupes=save_expenses_to_gsheet(client,parsed_exp)
                    if "parsed_df" in st.session_state:
                        write_pnl_sheet(client,st.session_state["parsed_df"],parsed_exp)
                st.success(f"✅ {added:,} expense rows saved. P&L sheet updated.")
                st.cache_data.clear()
        except Exception as e:
            st.error(f"Expense parse error: {e}")

    if "expense_df" not in st.session_state or st.session_state.get("expense_df",pd.DataFrame()).empty:
        loaded_exp=load_expenses_from_gsheet()
        if not loaded_exp.empty:
            st.session_state["expense_df"]=loaded_exp
            st.caption(f"📥 {len(loaded_exp)} expense records loaded from sheet")

    st.markdown("---")
    if "parsed_df" in st.session_state:
        df_all=st.session_state["parsed_df"]
    else:
        df_raw=load_from_gsheet()
        if df_raw.empty:
            st.info("No data yet. Upload a file above.")
            st.stop()
        df_all=enrich(df_raw)

    st.markdown("### 🔍 Filters")
    channels_avail=sorted(df_all["Channel"].unique())
    months_df=df_all[["Month_name","Month_sort"]].drop_duplicates().sort_values("Month_sort")
    months_avail=months_df["Month_name"].tolist()
    sel_channels=st.multiselect("Channels",channels_avail,default=channels_avail)
    sel_months=st.multiselect("Months",months_avail,default=months_avail)
    st.markdown("---")
    view=st.radio("View",["P&L Summary","Product P&L","Marketing Spend","Channel Deep-Dive","Month Trend","Channel Mix"])
    st.markdown("---")
    if st.button("🗑️ Clear Cache"):
        st.cache_data.clear(); st.success("Cache cleared.")
    if st.button("🔥 Clear Sheet & Re-upload",type="secondary"):
        with st.spinner("Clearing sheet..."):
            try:
                client=get_gsheet_client()
                ws=get_sheet(client).sheet1; ws.clear()
                st.cache_data.clear(); st.success("Sheet cleared. Re-upload your file.")
            except Exception as e: st.error(f"Error: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("<div class='header-bar'>📊 KENAZ PERFUMES — P&L DASHBOARD</div>",unsafe_allow_html=True)
df=df_all[df_all["Channel"].isin(sel_channels)&df_all["Month_name"].isin(sel_months)]
if df.empty:
    st.warning("No data for selected filters."); st.stop()

tot=df[["Net Sales","COGS","GM","CM1","CM2","Ad Spend","Quantity"]].sum()
nsv=tot["Net Sales"]

def kpi(col,label,val,sub=""):
    col.markdown(f"""<div class='kpi-card'><div class='kpi-label'>{label}</div><div class='kpi-value'>{val}</div><div class='kpi-sub'>{sub}</div></div>""",unsafe_allow_html=True)

exp_for_kpi=st.session_state.get("expense_df",pd.DataFrame())
brand_spend_tot=0.0
if not exp_for_kpi.empty:
    brand_spend_tot=float(exp_for_kpi[exp_for_kpi["Month"].isin(sel_months)]["Amount"].sum())
cm3_tot=float(tot["CM2"])-brand_spend_tot

if not exp_for_kpi.empty:
    c1,c2,c3,c4,c5,c6,c7=st.columns(7)
else:
    c1,c2,c3,c4,c5,c6=st.columns(6); c7=None

kpi(c1,"Net Sales",f"&#8377;{nsv/1e5:,.1f}L",f"{int(tot['Quantity']):,} units")
kpi(c2,"Gross Margin",f"&#8377;{tot['GM']/1e5:,.1f}L",P(safe_div(tot['GM'],nsv)*100)+" of NSV")
kpi(c3,"CM1",f"&#8377;{tot['CM1']/1e5:,.1f}L",P(safe_div(tot['CM1'],nsv)*100)+" of NSV")
kpi(c4,"CM2",f"&#8377;{tot['CM2']/1e5:,.1f}L",P(safe_div(tot['CM2'],nsv)*100)+" of NSV")
if c7:
    kpi(c5,"Brand Mktg",f"&#8377;{brand_spend_tot/1e5:,.1f}L",P(safe_div(brand_spend_tot,nsv)*100)+" of NSV")
    kpi(c6,"CM3",f"&#8377;{cm3_tot/1e5:,.1f}L",P(safe_div(cm3_tot,nsv)*100)+" of NSV")
    kpi(c7,"Ad Spend",f"&#8377;{tot['Ad Spend']/1e5:,.1f}L",P(safe_div(tot['Ad Spend'],nsv)*100)+" ACOS")
else:
    kpi(c5,"Ad Spend",f"&#8377;{tot['Ad Spend']/1e5:,.1f}L",P(safe_div(tot['Ad Spend'],nsv)*100)+" ACOS")
    kpi(c6,"COGS",f"&#8377;{tot['COGS']/1e5:,.1f}L",P(safe_div(tot['COGS'],nsv)*100)+" of NSV")

st.markdown("<br>",unsafe_allow_html=True)
month_order=(df[["Month_name","Month_sort"]].drop_duplicates().sort_values("Month_sort")["Month_name"].tolist())

if view=="P&L Summary":
    tab_options=["Total"]+sel_channels
    selected_tab=st.radio("View by",tab_options,horizontal=True,label_visibility="collapsed")
    df_view=df if selected_tab=="Total" else df[df["Channel"]==selected_tab]
    st.markdown(build_pnl_table(df_view,sel_months,sel_channels),unsafe_allow_html=True)
    try:
        xlsx_bytes=build_pnl_excel(df_view,sel_months)
        tab_label=selected_tab.replace(" ","_")
        st.download_button(label="⬇️ Download P&L as Excel",data=xlsx_bytes,
                           file_name=f"Kenaz_PL_{tab_label}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e: st.caption(f"Excel export error: {e}")

elif view=="Product P&L":
    st.subheader("Product-wise P&L")
    if "last_upload_bytes" not in st.session_state:
        st.info("Re-upload your file to see product-level P&L (session was reset)."); st.stop()
    with st.spinner("Parsing product data..."):
        if "sku_df_cache" not in st.session_state:
            st.session_state["sku_df_cache"]=parse_sku_data(st.session_state["last_upload_bytes"])
        sku_df=st.session_state["sku_df_cache"]
    all_models=sorted(sku_df["Model"].unique())
    all_months_sku=(sku_df[["Month_label","Month_sort"]].drop_duplicates().sort_values("Month_sort")["Month_label"].tolist())
    target_months={"Oct-25","Nov-25","Dec-25","Jan-26","Feb-26","Mar-26","Apr-26"}
    avail_months=[m for m in all_months_sku if m in target_months]
    col_f1,col_f2,col_f3=st.columns([2,2,2])
    with col_f1: sel_models=st.multiselect("Products",all_models,default=all_models,key="sku_models")
    with col_f2: sel_sku_months=st.multiselect("Months",avail_months,default=avail_months,key="sku_months")
    with col_f3: sel_sku_ch=st.multiselect("Channels",CHANNELS,default=CHANNELS,key="sku_ch")
    sku_f=sku_df[sku_df["Model"].isin(sel_models)&sku_df["Month_label"].isin(sel_sku_months)&sku_df["Channel"].isin(sel_sku_ch)]
    if sku_f.empty: st.warning("No data for selected filters."); st.stop()
    month_ord=(sku_f[["Month_label","Month_sort"]].drop_duplicates().sort_values("Month_sort")["Month_label"].tolist())
    grp=sku_f.groupby(["Model","Month_label","Month_sort"])[["Net Sales","Quantity","MRP Sales","COGS","Freight Inward","Wages","Commission","Payment Gateway","Shipping","Bulk Logistic","Packaging","Warehousing","Others","Ad Spend"]].sum().reset_index()
    def sku_metrics(g):
        ns=g["Net Sales"].sum(); cogs=g["COGS"].sum(); fw=(g["Freight Inward"]+g["Wages"]).sum()
        gm_v=ns-cogs-fw; cnl=(g["Commission"]+g["Payment Gateway"]+g["Shipping"]+g["Bulk Logistic"]+g["Packaging"]+g["Warehousing"]+g["Others"]).sum()
        cm1=gm_v-cnl; ads=g["Ad Spend"].sum(); cm2=cm1-ads
        return pd.Series({"Net Sales":ns,"COGS":cogs,"GM":gm_v,"CnL":cnl,"CM1":cm1,"Ad Spend":ads,"CM2":cm2,"Qty":g["Quantity"].sum(),"MRP Sales":g["MRP Sales"].sum()})
    summary=grp.groupby(["Model","Month_label","Month_sort"]).apply(sku_metrics).reset_index()
    summary["GM%"]=summary["GM"]/summary["Net Sales"].replace(0,np.nan)*100
    summary["CM1%"]=summary["CM1"]/summary["Net Sales"].replace(0,np.nan)*100
    summary["CM2%"]=summary["CM2"]/summary["Net Sales"].replace(0,np.nan)*100
    st.markdown("#### Net Sales by Product (INR Lacs)")
    ns_pivot=summary.pivot_table(index="Model",columns="Month_label",values="Net Sales",aggfunc="sum").reindex(columns=month_ord)
    ns_pivot["Total"]=ns_pivot.sum(axis=1); ns_pivot=ns_pivot.sort_values("Total",ascending=False)
    def style_lacs(val):
        try: f=float(val); return f"{f/1e5:,.2f}" if f!=0 else "-"
        except: return str(val)
    heads="".join(f"<th>{m}</th>" for m in month_ord)+"<th>Total</th>"; body=""
    for model in ns_pivot.index:
        cells="".join(f"<td>{style_lacs(ns_pivot.loc[model,m] if m in ns_pivot.columns else 0)}</td>" for m in month_ord)
        body+=f"<tr><td><b>{model}</b></td>{cells}<td><b>{style_lacs(ns_pivot.loc[model,'Total'])}</b></td></tr>"
    tot_cells="".join(f"<td><b>{style_lacs(ns_pivot[m].sum())}</b></td>" for m in month_ord)+f"<td><b>{style_lacs(ns_pivot['Total'].sum())}</b></td>"
    body+=f"<tr class='total-row'><td>Total</td>{tot_cells}</tr>"
    st.markdown(f"<div style='overflow-x:auto'><table class='pnl-table'><thead><tr><th>Product</th>{heads}</tr></thead><tbody>{body}</tbody></table></div>",unsafe_allow_html=True)
    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown("#### Product Deep-Dive P&L")
    sel_prod=st.selectbox("Select Product",sorted(sku_f["Model"].unique()),key="sku_prod")
    prod_df=sku_f[sku_f["Model"]==sel_prod]
    prod_monthly=(prod_df.groupby(["Month_label","Month_sort"])[["Net Sales","Quantity","MRP Sales","COGS","Freight Inward","Wages","Commission","Payment Gateway","Shipping","Bulk Logistic","Packaging","Warehousing","Others","Ad Spend"]].sum().reset_index().sort_values("Month_sort").set_index("Month_label"))
    prod_months=[m for m in month_ord if m in prod_monthly.index]
    def pv(col,m):
        if m not in prod_monthly.index: return 0.0
        val=prod_monthly.loc[m,col]
        if hasattr(val,"iloc"): val=val.iloc[0]
        return float(val)
    def mat_m(m): return pv("Net Sales",m)-pv("COGS",m)
    def fw_m(m): return pv("Freight Inward",m)+pv("Wages",m)
    def gm_m(m): return mat_m(m)-fw_m(m)
    def cnl_m(m): return (pv("Commission",m)+pv("Payment Gateway",m)+pv("Shipping",m)+pv("Bulk Logistic",m)+pv("Packaging",m)+pv("Warehousing",m)+pv("Others",m))
    def cm1_m(m): return gm_m(m)-cnl_m(m)
    def cm2_m(m): return cm1_m(m)-pv("Ad Spend",m)
    tot_ns_prod=sum(pv("Net Sales",m) for m in prod_months) or np.nan
    def prow(label,vals_fn,show_pct=False,is_total=False,inverse=False):
        cls="total-row" if is_total else ""; cells=""
        for m in prod_months:
            val=vals_fn(m); s=Lbold(val/1e5) if is_total else L(val/1e5); c=color_val(val,inverse)
            cells+=f"<td style='color:{c}'>{s}</td>"
        tv=sum(vals_fn(m) for m in prod_months); ts=Lbold(tv/1e5) if is_total else L(tv/1e5); tc=color_val(tv,inverse)
        row=f"<tr class='{cls}'><td>{label}</td>{cells}<td style='color:{tc}'>{ts}</td></tr>"
        if show_pct:
            pcells=""
            for m in prod_months:
                ns_m=pv("Net Sales",m) or np.nan; pval=safe_div(vals_fn(m),ns_m)*100; pc=color_val(pval,inverse)
                pcells+=f"<td style='color:{pc}'>{P(pval)}</td>"
            tot_pval=safe_div(tv,tot_ns_prod)*100; tpc=color_val(tot_pval,inverse)
            row+=f"<tr class='pct-row'><td></td>{pcells}<td style='color:{tpc}'>{P(tot_pval)}</td></tr>"
        return row
    ph="".join(f"<th>{m}</th>" for m in prod_months)+"<th>Total</th>"; pb=""; ncol=len(prod_months)+2
    pb+=prow("MRP Sales",lambda m:pv("MRP Sales",m))
    qty_cells="".join(f"<td>{int(pv('Quantity',m)):,}</td>" for m in prod_months)
    pb+=f"<tr><td>Quantity</td>{qty_cells}<td>{int(sum(pv('Quantity',m) for m in prod_months)):,}</td></tr>"
    pb+=f"<tr class='section-gap'>{'<td></td>'*ncol}</tr>"
    pb+=prow("Net Sales",lambda m:pv("Net Sales",m),is_total=True)
    pb+=f"<tr class='section-gap'>{'<td></td>'*ncol}</tr>"
    pb+=prow("Less: COGS",lambda m:pv("COGS",m),show_pct=True,inverse=True)
    pb+=prow("Material Margins",lambda m:mat_m(m),show_pct=True,is_total=True)
    pb+=f"<tr class='section-gap'>{'<td></td>'*ncol}</tr>"
    pb+=prow("Less: Freight Inwards",lambda m:pv("Freight Inward",m),inverse=True)
    pb+=prow("Less: Wages - Fixed",lambda m:pv("Wages",m),inverse=True)
    pb+=prow("Freight & Wages Total",lambda m:fw_m(m),show_pct=True,inverse=True)
    pb+=prow("Gross Margin",lambda m:gm_m(m),show_pct=True,is_total=True)
    pb+=f"<tr class='section-gap'>{'<td></td>'*ncol}</tr>"
    pb+=prow("Less: Commission",lambda m:pv("Commission",m),inverse=True)
    pb+=prow("Less: Payment GW",lambda m:pv("Payment Gateway",m),inverse=True)
    pb+=prow("Less: Shipping",lambda m:pv("Shipping",m),inverse=True)
    pb+=prow("Less: Bulk Logistic",lambda m:pv("Bulk Logistic",m),inverse=True)
    pb+=prow("Less: Packaging",lambda m:pv("Packaging",m),inverse=True)
    pb+=prow("Less: Warehousing",lambda m:pv("Warehousing",m),inverse=True)
    pb+=prow("Less: Others",lambda m:pv("Others",m),inverse=True)
    pb+=prow("C&L Total",lambda m:cnl_m(m),show_pct=True,inverse=True)
    pb+=f"<tr class='section-gap'>{'<td></td>'*ncol}</tr>"
    pb+=prow("CM1",lambda m:cm1_m(m),show_pct=True,is_total=True)
    pb+=f"<tr class='section-gap'>{'<td></td>'*ncol}</tr>"
    pb+=prow("Less: Ad Spend",lambda m:pv("Ad Spend",m),show_pct=True,inverse=True)
    pb+=prow("CM2",lambda m:cm2_m(m),show_pct=True,is_total=True)
    st.markdown(f"<div style='overflow-x:auto'><table class='pnl-table'><thead><tr><th>{sel_prod} (INR Lacs)</th>{ph}</tr></thead><tbody>{pb}</tbody></table></div>",unsafe_allow_html=True)
    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown("#### CM2 Trend by Product")
    cm2_data=summary[summary["Month_label"].isin(prod_months)].copy()
    fig=px.bar(cm2_data.sort_values("Month_sort"),x="Month_label",y="CM2",color="Model",barmode="group",
               color_discrete_sequence=[GOLD,"#e67e22","#4fc3f7","#81c784","#ce93d8","#ef9a9a","#fff176","#b0bec5","#80cbc4","#ffcc80"])
    fig.update_layout(template="plotly_dark",plot_bgcolor=DARK,paper_bgcolor=DARK,font=dict(color="#aaa"),
                      yaxis_title="INR",legend=dict(orientation="h"),margin=dict(t=20,b=30,l=10,r=10),height=380,
                      xaxis=dict(gridcolor="#333",categoryorder="array",categoryarray=month_ord),yaxis=dict(gridcolor="#333"))
    st.plotly_chart(fig,use_container_width=True)

elif view=="Marketing Spend":
    st.subheader("Marketing & Influencer Spend — MoM")
    if "expense_df" not in st.session_state or st.session_state["expense_df"].empty:
        st.info("Upload the Expense Sheet (.xlsx) in the sidebar to see this view."); st.stop()
    exp=st.session_state["expense_df"]
    MONTH_ORD=["Sep-25","Oct-25","Nov-25","Dec-25","Jan-26","Feb-26","Mar-26","Apr-26"]
    avail_months=[m for m in MONTH_ORD if m in exp["Month"].unique()]
    inf_tot=exp[exp["Section"]=="Influencer Spend"]["Amount"].sum()
    mkt_tot=exp[exp["Section"]=="Marketing Spend"]["Amount"].sum()
    k1,k2,k3=st.columns(3)
    def ekpi(col,label,val): col.markdown(f"<div class='kpi-card'><div class='kpi-label'>{label}</div><div class='kpi-value'>&#8377;{val/1e5:,.2f}L</div></div>",unsafe_allow_html=True)
    ekpi(k1,"Influencer Spend (Total)",inf_tot); ekpi(k2,"Marketing Spend (Total)",mkt_tot); ekpi(k3,"Combined Total",inf_tot+mkt_tot)
    st.markdown("<br>",unsafe_allow_html=True)
    st.markdown("#### MoM Breakdown (INR Lacs)")
    mom=exp.groupby(["Section","Month"])["Amount"].sum().unstack(fill_value=0)
    mom=mom.reindex(columns=[m for m in avail_months if m in mom.columns]); mom["Total"]=mom.sum(axis=1)
    heads="".join(f"<th>{m}</th>" for m in mom.columns); body=""
    for section in mom.index:
        cells="".join(f"<td style='color:{GOLD if col=='Total' else '#ddd'}'>{mom.loc[section,col]/1e5:,.2f}</td>" for col in mom.columns)
        body+=f"<tr><td><b>{section}</b></td>{cells}</tr>"
    tot_cells="".join(f"<td style='color:{GOLD}'><b>{mom[col].sum()/1e5:,.2f}</b></td>" for col in mom.columns)
    body+=f"<tr class='total-row'><td>Grand Total</td>{tot_cells}</tr>"
    st.markdown(f"<div style='overflow-x:auto'><table class='pnl-table'><thead><tr><th>Section</th>{heads}</tr></thead><tbody>{body}</tbody></table></div>",unsafe_allow_html=True)
    st.markdown("<br>",unsafe_allow_html=True)
    chart_data=exp.groupby(["Section","Month"])["Amount"].sum().reset_index()
    chart_data=chart_data[chart_data["Month"].isin(avail_months)]
    chart_data["Month"]=pd.Categorical(chart_data["Month"],categories=avail_months,ordered=True)
    chart_data=chart_data.sort_values("Month")
    fig=go.Figure()
    for section,color in [("Influencer Spend",GOLD),("Marketing Spend","#e67e22")]:
        sd=chart_data[chart_data["Section"]==section]
        fig.add_trace(go.Bar(name=section,x=sd["Month"],y=sd["Amount"]/1e5,marker_color=color))
    fig.update_layout(barmode="stack",template="plotly_dark",plot_bgcolor=DARK,paper_bgcolor=DARK,
                      font=dict(color="#aaa"),yaxis_title="INR Lacs",legend=dict(orientation="h"),height=380,
                      margin=dict(t=20,b=30,l=10,r=10),xaxis=dict(gridcolor="#333"),yaxis=dict(gridcolor="#333"))
    st.plotly_chart(fig,use_container_width=True)
    st.markdown("#### Vendor Detail")
    col_sec,col_mon=st.columns(2)
    with col_sec: sec_filter=st.selectbox("Section",["All","Influencer Spend","Marketing Spend"])
    with col_mon: mon_filter=st.selectbox("Month",["All"]+avail_months)
    detail=exp.copy()
    if sec_filter!="All": detail=detail[detail["Section"]==sec_filter]
    if mon_filter!="All": detail=detail[detail["Month"]==mon_filter]
    detail=detail[detail["Amount"]>0].sort_values("Amount",ascending=False)
    vend_pivot=detail.groupby(["Vendor","Nature","Section"])["Amount"].sum().reset_index().sort_values("Amount",ascending=False)
    body2="".join(f"<tr><td>{r['Vendor']}</td><td style='color:#aaa;font-size:12px'>{r['Nature']}</td><td style='color:#888;font-size:12px'>{r['Section']}</td><td style='color:{GOLD}'>{r['Amount']/1e5:,.2f}L</td></tr>" for _,r in vend_pivot.iterrows())
    if not body2: body2=f"<tr><td colspan='4' style='text-align:center;color:#555'>No data</td></tr>"
    body2+=f"<tr class='total-row'><td><b>Total</b></td><td></td><td></td><td><b>{vend_pivot['Amount'].sum()/1e5:,.2f}L</b></td></tr>"
    st.markdown(f"<div style='overflow-x:auto'><table class='pnl-table'><thead><tr><th>Vendor</th><th>Nature</th><th>Section</th><th>Amount</th></tr></thead><tbody>{body2}</tbody></table></div>",unsafe_allow_html=True)

elif view=="Channel Deep-Dive":
    st.subheader("Channel Deep-Dive")
    ch_sel=st.selectbox("Channel",sel_channels)
    dch=df[df["Channel"]==ch_sel]
    pivot=(dch.groupby("Month_name").agg(Net_Sales=("Net Sales","sum"),GM=("GM","sum"),CM1=("CM1","sum"),CM2=("CM2","sum"),Ad_Spend=("Ad Spend","sum")).reindex([m for m in month_order if m in dch["Month_name"].values]).reset_index())
    fig=go.Figure()
    fig.add_trace(go.Bar(name="CM2",x=pivot["Month_name"],y=pivot["CM2"]/1e5,marker_color=GOLD))
    fig.add_trace(go.Bar(name="Ad Spend",x=pivot["Month_name"],y=pivot["Ad_Spend"]/1e5,marker_color="#e67e22"))
    fig.add_trace(go.Scatter(name="Net Sales",x=pivot["Month_name"],y=pivot["Net_Sales"]/1e5,mode="lines+markers",line=dict(color="#4fc3f7",width=2)))
    fig.update_layout(barmode="stack",template="plotly_dark",plot_bgcolor=DARK,paper_bgcolor=DARK,
                      font=dict(color="#aaa"),yaxis_title="INR Lakhs",legend=dict(orientation="h"),
                      margin=dict(t=30,b=30,l=10,r=10),height=360,
                      xaxis=dict(gridcolor="#333"),yaxis=dict(gridcolor="#333"))
    st.plotly_chart(fig,use_container_width=True)
    st.markdown(build_pnl_table(dch,sel_months,[ch_sel]),unsafe_allow_html=True)

elif view=="Month Trend":
    st.subheader("Monthly Trend")
    metric=st.selectbox("Metric",["Net Sales","GM","CM1","CM2","Ad Spend","COGS"])
    monthly=(df.groupby(["Month_name","Month_sort"])[metric].sum().reset_index().sort_values("Month_sort"))
    fig=go.Figure(go.Bar(x=monthly["Month_name"],y=monthly[metric]/1e5,marker_color=GOLD,name=metric))
    fig.update_layout(template="plotly_dark",plot_bgcolor=DARK,paper_bgcolor=DARK,font=dict(color="#aaa"),
                      yaxis_title="INR Lakhs",margin=dict(t=30,b=30,l=10,r=10),height=380,
                      xaxis=dict(gridcolor="#333"),yaxis=dict(gridcolor="#333"))
    st.plotly_chart(fig,use_container_width=True)
    ch_monthly=(df.groupby(["Channel","Month_name","Month_sort"])[metric].sum().reset_index().sort_values("Month_sort"))
    fig2=px.line(ch_monthly,x="Month_name",y=metric,color="Channel",
                 color_discrete_sequence=[GOLD,"#e67e22","#4fc3f7","#81c784","#ce93d8"],
                 category_orders={"Month_name":month_order})
    fig2.update_layout(template="plotly_dark",plot_bgcolor=DARK,paper_bgcolor=DARK,font=dict(color="#aaa"),
                       margin=dict(t=20,b=30,l=10,r=10),height=350,
                       xaxis=dict(gridcolor="#333"),yaxis=dict(gridcolor="#333"),legend=dict(orientation="h"))
    st.plotly_chart(fig2,use_container_width=True)

elif view=="Channel Mix":
    st.subheader("Channel Mix")
    ch_agg=df.groupby("Channel")[["Net Sales","CM1","CM2","Ad Spend","Quantity"]].sum().reset_index()
    COLORS=[GOLD,"#e67e22","#4fc3f7","#81c784","#ce93d8"]
    col1,col2=st.columns(2)
    with col1:
        fig=px.pie(ch_agg,names="Channel",values="Net Sales",title="Revenue Split",color_discrete_sequence=COLORS)
        fig.update_layout(template="plotly_dark",paper_bgcolor=DARK,font=dict(color="#aaa"),margin=dict(t=40,b=10),height=340)
        st.plotly_chart(fig,use_container_width=True)
    with col2:
        fig=px.pie(ch_agg,names="Channel",values="CM2",title="CM2 Split",color_discrete_sequence=COLORS)
        fig.update_layout(template="plotly_dark",paper_bgcolor=DARK,font=dict(color="#aaa"),margin=dict(t=40,b=10),height=340)
        st.plotly_chart(fig,use_container_width=True)
    ch_agg["CM1%"]=ch_agg["CM1"]/ch_agg["Net Sales"].replace(0,np.nan)*100
    ch_agg["CM2%"]=ch_agg["CM2"]/ch_agg["Net Sales"].replace(0,np.nan)*100
    ch_agg=ch_agg.sort_values("Net Sales",ascending=True)
    fig3=go.Figure()
    fig3.add_trace(go.Bar(name="CM1%",x=ch_agg["CM1%"],y=ch_agg["Channel"],orientation="h",marker_color=GOLD))
    fig3.add_trace(go.Bar(name="CM2%",x=ch_agg["CM2%"],y=ch_agg["Channel"],orientation="h",marker_color="#e67e22"))
    fig3.update_layout(barmode="group",template="plotly_dark",plot_bgcolor=DARK,paper_bgcolor=DARK,
                       font=dict(color="#aaa"),xaxis_title="% of Net Sales",
                       margin=dict(t=20,b=30,l=10,r=10),height=320,legend=dict(orientation="h"))
    st.plotly_chart(fig3,use_container_width=True)

st.markdown(f"""<hr style='border-color:#333;margin-top:40px'>
<div style='text-align:center;color:#555;font-size:11px'>Kenaz Perfumes · One Guardian Brands · FY 25-26 P&amp;L Analytics</div>""",unsafe_allow_html=True)
